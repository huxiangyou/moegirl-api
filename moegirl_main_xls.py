"""
Python 3.8.0
检查页面、讨论、重定向是否被编辑、移动、重定向、删除、创建
大约需要2分钟

Jan 17, 2020
Hu Xiangyou
"""
print(__file__.rsplit("\\",1)[1])
print(__doc__)
import requests
from urllib import parse
import time
import openpyxl
import second2days

S_M="[页]"
S_T="[论]"
S_R="[重]"
S_REDI="[重定向]"
S_DELE="[删除]"
S_MOVE="[移动]"
S_NEWW="[新]"
S_EDIT="[编辑]"
S_TOOO="->"
S_NOTO="-≯"
S_PID="[PID]"
S_TID="[Talk]"
S_RID="[Rev]"
S_TIL="[Title]"
S_RED="[R]"

filepath="PAGELIST.xlsx"

url="https://zh.moegirl.org/api.php"

params={'action':'query','format':'json','prop':'info|redirects','curtimestamp':1,'indexpageids':1,'rdlimit':500,'inprop':'talkid','pageids':''}

Workbook:type=openpyxl.workbook.workbook.Workbook
Worksheet:type=openpyxl.worksheet.worksheet.Worksheet
Cell:type=openpyxl.cell.cell.Cell

b:Workbook=openpyxl.load_workbook(filepath)
sMain:Worksheet=b.worksheets[0]
sTemp:Worksheet=b.worksheets[1]
sCate:Worksheet=b.worksheets[2]
sheets:Worksheet=b.worksheets

n_all=0
n_m_moved=0
n_m_redirects=0
n_m_deleted=0
n_m_edited=0
n_t_new=0
n_t_deleted=0
n_r_all=0
n_r_new=0
n_r_deleted=0

def devide(l:list,n:int)->list:
	for i in range(0,len(l),n):
		yield l[i:i+n]

def findRowIndex(sheet:Worksheet,column:int,value)->int:
	return list(ri.value for ri in list(sheet.columns)[column]).index(value)+1

def findCell(sheet:Worksheet,pageid:int,column:int)->Cell:
	return sheet.cell(findRowIndex(sheet,1,pageid),column)

def changeCell(sheet:Worksheet,pageid:int,column:int,value)->Cell:
	return sheet.cell(findRowIndex(sheet,1,pageid),column,str(value))

def pageid2revid(sheet:Worksheet,pageid:int)->str:
	value=findCell(sheet,pageid,3).value
	if value:
		return int(value)
	else:
		return 0

def pageid2title(sheet:Worksheet,pageid:int)->str:
	return findCell(sheet,pageid,4).value

def pageid2talkid(sheet:Worksheet,pageid:int)->int:
	value=findCell(sheet,pageid,5).value
	if value:
		return int(value)
	else:
		return 0

def pageid2redirectSet(sheet:Worksheet,pageid:int)->set:
	value=findCell(sheet,pageid,6).value
	if value and eval(value):
		return eval(value)
	else:
		return set()

def isResponceOK(json:dict)->bool:
	if 'batchcomplete' in json:
		return True
	else:
		print("Responce is not complete.")
		return False

def pageDeleted(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_m_deleted
	oldTitle:str=pageid2title(sheet,pageid)
	if 'missing' in json['query']['pages'][str(pageid)]:
		print(S_M+S_DELE,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,sep="")
		pass
		n_m_deleted+=1
		return True
	else:
		return False

def pageMoved(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_m_moved
	oldTitle:str=pageid2title(sheet,pageid)
	newTitle=json['query']['pages'][str(pageid)]['title']
	if newTitle!=oldTitle:
		print(S_M+S_MOVE,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,S_TOOO,newTitle,sep="")
		changeCell(sheet,pageid,4,newTitle)
		n_m_moved+=1
		return True
	else:
		return False

def pageRedirect(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_m_redirects
	oldTitle:str=pageid2title(sheet,pageid)
	if 'redirect' in json['query']['pages'][str(pageid)]:
		print(S_M+S_REDI,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,sep="")
		pass
		n_m_redirects+=1
		return True
	else:
		return False

def pageEdited(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_m_edited
	oldrevid:str=pageid2revid(sheet,pageid)
	lastrevid=json['query']['pages'][str(pageid)]['lastrevid']
	oldTitle:str=pageid2title(sheet,pageid)
	if oldrevid!=lastrevid:
		if oldrevid=="0":
			print(S_M+S_NEWW,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_RID,lastrevid,sep="")
		else:
			print(S_M+S_EDIT,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_RID,oldrevid,S_TOOO,lastrevid,sep="")
		changeCell(sheet,pageid,3,lastrevid)
		n_m_edited+=1
		return True
	else:
		return False

def talkCreated(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_t_new
	oldTitle:str=pageid2title(sheet,pageid)
	if 'talkid' in json['query']['pages'][str(pageid)]:
		talkid=json['query']['pages'][str(pageid)]['talkid']
		print(S_T+S_NEWW,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_TID,talkid,sep="")
		changeCell(sheet,pageid,5,talkid)
		n_t_new+=1
		return True
	else:
		return False

def talkDeleted(json:dict,sheet:Worksheet,pageid:int)->bool:
	global n_t_deleted
	oldTitle:str=pageid2title(sheet,pageid)
	talkid:int=pageid2talkid(sheet,pageid)
	if 'talkid' not in json['query']['pages'][str(pageid)]:
		print(S_T+S_DELE,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_TID,talkid,sep="")
		changeCell(sheet,pageid,5,"")
		n_t_deleted+=1
		return True
	else:
		return False

def getRedirectSet(json:dict,pageid:int)->set:
	if 'redirects' in json['query']['pages'][str(pageid)]:
		return set((r['title'] for r in json['query']['pages'][str(pageid)]['redirects']))
	else:
		return set()

def redirectCreated(sheet:Worksheet,pageid:int,newSet:set)->bool:
	global n_r_new
	oldTitle:str=pageid2title(sheet,pageid)
	oldSet:set=pageid2redirectSet(sheet,pageid)
	if newSet-oldSet:
		print(S_R+S_NEWW,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_RED,newSet-oldSet,sep="")
		if newSet:
			changeCell(sheet,pageid,6,newSet)
		else:
			changeCell(sheet,pageid,6,"")
		n_r_new+=len(newSet-oldSet)
		return newSet-oldSet
	else:
		return set()

def redirectDeleted(sheet:Worksheet,pageid:int,newSet:set)->bool:
	global n_r_deleted
	oldTitle:str=pageid2title(sheet,pageid)
	oldSet:set=pageid2redirectSet(sheet,pageid)
	if oldSet-newSet:
		print(S_R+S_DELE,"\t",S_PID,pageid,"\t",S_TIL,oldTitle,"\t",S_RED,oldSet-newSet,sep="")
		if newSet:
			changeCell(sheet,pageid,6,newSet)
		else:
			changeCell(sheet,pageid,6,"")
		n_r_deleted+=len(oldSet-newSet)
		return oldSet-newSet
	else:
		return set()

def showPageAll(json:dict,pageid:int):
	pageJson=json['query']['pages'][str(pageid)]
	pageid=pageJson['pageid'] if 'pageid' in pageJson else ""
	lastrevid=pageJson['lastrevid'] if 'lastrevid' in pageJson else ""
	title=pageJson['title'] if 'title' in pageJson else ""
	touched=pageJson['touched'] if 'touched' in pageJson else ""
	length=pageJson['length'] if 'length' in pageJson else ""
	talkid=pageJson['talkid'] if 'talkid' in pageJson else ""
	redirects=getRedirectSet(json,pageid)
	redirects=redirects if redirects else ""
	print("-","\t",S_PID,pageid,"\t",S_RID,lastrevid,"\t",touched,"\t",S_TIL,title,sep="")
	if talkid:
		print("\t",S_TID,talkid,sep="",end="")
	if redirects:
		print("\t",S_RED,redirects,sep="")

input("Please close the Excel file first. Enter to start.")
start_time=time.time()
print("started at",time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(start_time)))

for sheet in sheets:
	pageidlist:list=[c.value for c in list(sheet.columns)[1]]
	n_all+=len(pageidlist)
	for pageidlist_d in devide(pageidlist,50):
		l_param:str="|".join(str(i) for i in pageidlist_d)
		while True:
			try:
				a=requests.get(url,params={**params,**{'pageids':l_param}},timeout=(10,30))
			except:
				print(".",end="")
				continue
			else:
				break
		for pageid in pageidlist_d:
			json:dict=a.json()
			if not isResponceOK(json):
				break

			#showPageAll(json,pageid)

			#is deleted
			if not pageDeleted(json,sheet,pageid):

				#is moved
				pageMoved(json,sheet,pageid)

				#is redirect
				pageRedirect(json,sheet,pageid)

				#is edited
				pageEdited(json,sheet,pageid)

				#is talk created
				if not pageid2talkid(sheet,pageid):
					talkCreated(json,sheet,pageid)

				#is talk deleted
				if pageid2talkid(sheet,pageid):
					talkDeleted(json,sheet,pageid)

				#redirects
				newSet=getRedirectSet(json,pageid)
				n_r_all+=len(newSet)
				redirectCreated(sheet,pageid,newSet)
				redirectDeleted(sheet,pageid,newSet)
while True:
	try:
		b.save(filepath)
	except:
		input("Please close the file to continue saving the data.")
	else:
		break
print()
end_time=time.time()
print("ended   at",time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(end_time)))
print(second2days.main(end_time-start_time),"spent.")

print(n_all,"intries checked.")
print(n_m_moved,"moved.",n_m_redirects,"redirected.",n_m_deleted,"deleted.",n_m_edited,"edited.")
print("Talk pages:",n_t_new,"newly created.",n_t_deleted,"deleted.")
print("Redirects: ",n_r_new,"newly created.",n_r_deleted,"deleted.",n_r_all,"found.")

input("Done.")